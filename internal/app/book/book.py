from io import BytesIO
from typing import Optional

import openpyxl

from internal.adapters.repository.repository import BookRepoIface
from internal.adapters.storage.gbacket.book_storage import StorageBook
from internal.app.app import BookAppIface
from internal.app.book.dto.book import BookListSchema, BookSchema, BookFilter, BookUpdate, BookBatchUpdate
from internal.ports.http.book.dto.book_create import BookCreateBody
from pkg.server.aiohttp.errors import ErrNotFount, ErrForbidden


class BookApp(BookAppIface):
    def __init__(self, repo: BookRepoIface, storage: StorageBook):
        self.repo = repo
        self.storage = storage

    async def get_book(self, book_id: int) -> Optional[BookSchema]:
        book = await self.repo.get_book(book_id)
        return BookSchema.from_orm(book)

    async def get_books(self, filters: BookFilter) -> BookListSchema:
        books = await self.repo.get_books(filters)
        return BookListSchema([BookSchema.from_orm(book) for book in books])

    async def create_book(self, book: BookCreateBody) -> int:
        author = await self.repo.get_author(book.author_id)
        if not author:
            raise ErrNotFount("Author with this id not fount")

        book_data = book.model_dump()
        author_id = book_data.pop("author_id")

        # book_data["file_url"] = file_url
        book_id = await self.repo.create_book(book_data)

        await self.repo.attach_book_to_author(
            author_id=author_id,
            book_id=book_id
        )

        return book_id

    async def upload_book_file(self, book_id: int, file_bytes: bytes) -> None:
        book = await self.repo.get_book(book_id)
        if not book:
            raise ErrNotFount("Author with this id not fount")

        file_name = f"{book.title}-{book.id}.pdf"

        file_url = await self.storage.upload_book(file_bytes, file_name)

        await self.repo.update_book(book_id, BookUpdate(file_url=file_url))

    async def download_book_file(self, book_id: int, read_only: Optional[bool] = False) -> (bytes, str):
        book = await self.repo.get_book(book_id)
        if not book:
            raise ErrNotFount("Book with this id not fount")

        if read_only and book.restricted:
            raise ErrForbidden("Book not allowed for stream")

        file = await self.storage.download_book(book.file_url)
        if not file:
            raise ErrNotFount("File with this id not fount")

        return file, book.file_url

    async def upload_book_restrictions(self, restrictions_file: bytes) -> None:
        file_io = BytesIO(restrictions_file)

        wb = openpyxl.load_workbook(file_io, read_only=True)

        name = wb["name"]
        author = wb["author"]

        title_list = []
        author_list = []

        for row in name.iter_rows(min_row=2, min_col=2):
            if title := row[0].value:
                title_list.append(str(title))

        for row in author.iter_rows(min_row=2, min_col=2):
            if author := row[0].value:
                author_list.append(str(author))

        await self.repo.restrict_books(
            BookBatchUpdate(
                titles_list=title_list,
                authors_list=author_list)
        )
